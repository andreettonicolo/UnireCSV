import openpyxl
from openpyxl.styles import PatternFill, Border, Side
import pandas as pd

def convert_to_xlsx(merged_data, output_xlsx_file):
    with pd.ExcelWriter(output_xlsx_file, engine='openpyxl') as writer:
        merged_data.to_excel(writer, sheet_name='Sheet1', index=False)

        # Rimuovi colonne vuote anche dal file XLSX
        for sheet in writer.sheets.values():
            for col_cells in sheet.columns:
                if all(cell.value is None for cell in col_cells):
                    sheet.delete_cols(col_cells[0].col_idx, len(col_cells))

    wb = openpyxl.load_workbook(output_xlsx_file)
    sheet = wb['Sheet1']

    # Calcola la larghezza delle colonne
    column_widths = []
    for col in merged_data.columns:
        # Trova la lunghezza massima nella colonna (comprensiva dell'intestazione)
        column_length = max(merged_data[col].astype(str).apply(len).max(), len(str(col)))
        # Aggiungi un margine aggiuntivo di 2 per garantire che il testo sia leggibile
        column_widths.append(column_length + 2)

    # Applica la larghezza alle colonne
    for i, width in enumerate(column_widths, start=1):
        col_letter = openpyxl.utils.get_column_letter(i)
        sheet.column_dimensions[col_letter].width = width

    # Crea una scala di due tonalità di verde pastello
    green_gradient = ['DFF0D8', 'A9DFBF']

    # Applica il gradiente di colore e i bordi alle celle della tabella in alternanza
    for row in range(2, sheet.max_row + 1):
        gradient_index = (row - 2) % len(green_gradient)  # Alterna tra le tonalità di verde pastello
        fill_color = green_gradient[gradient_index]

        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

            # Aggiungi bordi interni
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = border

    # Salva le modifiche nel file XLSX
    wb.save(output_xlsx_file)
