import openpyxl

# Percorso del file merged.xlsx
merged_xlsx_file = '../xlsx_finale/merged.xlsx'

# Apri il file XLSX
wb = openpyxl.load_workbook(merged_xlsx_file)

# Seleziona il foglio di lavoro
sheet = wb['Sheet1']

# Trova l'indice delle colonne da eliminare
columns_to_delete = ['Network Agent is installed', 'Real-time protection status', 'Created']
columns_to_delete_indices = []

for col in range(1, sheet.max_column + 1):
    col_name = sheet.cell(row=1, column=col).value
    if col_name in columns_to_delete:
        columns_to_delete_indices.append(col)

# Elimina le colonne
for col_index in sorted(columns_to_delete_indices, reverse=True):
    sheet.delete_cols(col_index)

# Trova l'indice della colonna 'Last connected to Administration Server'
last_connected_col_index = None
for col in range(1, sheet.max_column + 1):
    if sheet.cell(row=1, column=col).value == 'Last connected to Administration Server':
        last_connected_col_index = col
        break

# Se l'indice della colonna è stato trovato, elabora le celle
if last_connected_col_index:
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=last_connected_col_index)
        cell_value = cell.value

        # Controlla se "minutes" o "hour" è presente nella cella
        if isinstance(cell_value, str) and ('minutes' in cell_value or 'hour' in cell_value):
            # Elimina il testo "minutes" o "hour" dalla cella
            cell.value = "Online recently"

# Salva le modifiche nel file XLSX
wb.save(merged_xlsx_file)

print('Elaborazione del file XLSX completata. Le modifiche sono state apportate.')
